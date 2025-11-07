import csv
from django.http import HttpResponse
from django.views.generic import ListView, CreateView, UpdateView
from django.urls import reverse_lazy
from .models import BastRecordModel
from cl_seiscomp.models import CsRecordModel
from core.models import Kelompok
from .forms import BastRecordForm
import requests, openpyxl, datetime, os
import pandas as pd
from django.http import JsonResponse, HttpResponse
from core.models import Operator
from io import StringIO
from openpyxl.utils.dataframe import dataframe_to_rows
from django.views import View
from django.shortcuts import redirect
from django.forms.models import model_to_dict

def bastrecord_list_api(request, counts=0):
    if counts > 0:
        records = BastRecordModel.objects.all().order_by('-bast_id').select_related('spv')[:counts]
    else:
        records = BastRecordModel.objects.all().order_by('-bast_id').select_related('spv')
        
    # Serialize with related supervisor name
    data = []
    for record in records:
        record_dict = model_to_dict(record)
        # Add supervisor name if exists
        if record.spv:
            record_dict['spv_name'] = record.spv.name
        else:
            record_dict['spv_name'] = ''
        data.append(record_dict)
    return JsonResponse(data, safe=False)

class BastRecordListView(ListView):
    model = BastRecordModel
    template_name = 'bast/bastrecord_list.html'
    context_object_name = 'bastrecords'

class BastRecordCreateView(CreateView):
    model = BastRecordModel
    form_class = BastRecordForm
    template_name = 'bast/bastrecord_form.html'
    success_url = reverse_lazy('bast:bastrecord_list')

class BastRecordUpdateView(UpdateView):
    model = BastRecordModel
    form_class = BastRecordForm
    template_name = 'bast/bastrecord_form.html'
    success_url = reverse_lazy('bast:bastrecord_list')

    def form_valid(self, form):
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        record = self.get_object()
        context['existing_data'] = record.events
        context['existing_member_data'] = record.member
        return context

    def get_member_data(self, **kwargs):
        context = super().get_member_data(**kwargs)
        record = self.get_object()
        context['existing_member_data'] = record.member
        return context

class BastAllRecordsView(ListView):
    model = BastRecordModel
    template_name = 'bast/bast_all_records.html'
    context_object_name = 'bastrecords'

class BastRecordDeleteDirectView(View):
    def post(self, request, pk, *args, **kwargs):
        try:
            record = BastRecordModel.objects.get(pk=pk)
            record.delete()
            return redirect('bast:bastrecord_list')
        except BastRecordModel.DoesNotExist:
            return HttpResponse(status=404)

# Functions
def clean_index3(data, start_datetime='2024-12-11 13:00:00', end_datetime='2024-12-11 19:00:00'):
    text = data.decode('utf-8')

    lines = text.split('\n')
    processed_lines = []
    for i, line in enumerate(lines):
        if i not in [0, 1, 3]:
            line = '|'.join(part.strip() for part in line.split('|'))
            processed_lines.append(line)

    df = pd.DataFrame([x.split('|') for x in processed_lines[1:]], columns=processed_lines[0].split('|'))
    df['Origin Time (GMT)'] = pd.to_datetime(df['Origin Time (GMT)'], format='%Y-%m-%d %H:%M:%S')

    def select_data_by_datetime_range(df, start_datetime, end_datetime):
        mask = (df['Origin Time (GMT)'] >= start_datetime) & (df['Origin Time (GMT)'] <= end_datetime)
        return df.loc[mask]

    df_selected = select_data_by_datetime_range(df, start_datetime, end_datetime)

    # sort the df_selected by 'Origin Time (GMT)'
    df_selected = df_selected.sort_values(by='Origin Time (GMT)')

    # divide the 'Origin Time (GMT)' column into 'Date' and 'OT (UTC)' columns, and put them in the first two columns, remove the 'Origin Time (GMT)' column
    df_selected['Date'] = df_selected['Origin Time (GMT)'].dt.date
    df_selected['OT (UTC)'] = df_selected['Origin Time (GMT)'].dt.time
    df_selected = df_selected[['Date', 'OT (UTC)'] + [col for col in df_selected.columns if col != 'Origin Time (GMT)']]
    df_selected = df_selected.reset_index(drop=True)

    # sort the columns to be 'Date', 'OT (UTC)', 'Lat', 'Long', 'Mag', 'D(Km)', 'Phase', 'RMS', 'Az.Gap', 'Region', but first turn the respective column names into the desired ones
    df_selected = df_selected.rename(columns={'Lon': 'Long', 'Depth': 'D(Km)', 'cntP': 'Phase', 'AZgap': 'Az. Gap', 'Remarks': 'Region'})
    df_selected = df_selected[['Date', 'OT (UTC)', 'Lat', 'Long', 'D(Km)', 'Mag', 'TypeMag', 'Region']]
    df_selected = df_selected.reset_index(drop=True)

    # add numbering to the first column
    df_selected.insert(0, 'No', range(1, len(df_selected) + 1))

    # Check for duplicate columns
    df_selected = df_selected.loc[:, ~df_selected.columns.duplicated()]

    # add MMI, terkirim M>5, and terkirim M>5 columns with empty values
    df_selected['MMI'] = ''
    df_selected['Dis. PGN'] = ''
    df_selected['Selisih PGN'] = ''
    df_selected['Dis. PGR'] = '' 
    df_selected['Selisih PGR'] = '' 
    
    return df_selected

def fetch_data(request, start_datetime='2024-12-11 13:00:00', end_datetime='2024-12-11 19:00:00'):
    url = "http://202.90.198.41/index3.txt"
    response = requests.get(url)

    if response.status_code == 200:
        data = clean_index3(response.content, start_datetime, end_datetime)
        csv_data = data.to_csv(index=False)
        table_data = data.to_dict(orient='records')
        return JsonResponse({'csv': csv_data, 'table_data': table_data})
    else:
        return JsonResponse({'error': 'Failed to fetch data'}, status=500)

def get_nip(request, operator_id):
    try:
        operator = Operator.objects.get(id=operator_id)
        return JsonResponse({'nip': operator.NIP})
    except Operator.DoesNotExist:
        return JsonResponse({'error': 'Operator not found'}, status=404)

def get_member_data(request, kelompok):
    try:
        kelompok = Kelompok.objects.get(name=kelompok)
        member_data = kelompok.member
        member_pks = [int(member.strip()) for member in member_data.split(',')]
        member_names = [Operator.objects.get(pk=pk).name for pk in member_pks]
        return JsonResponse({'member_data': member_names})
    except Kelompok.DoesNotExist:
        return JsonResponse({'error': 'Kelompok not found'}, status=404)
    except Operator.DoesNotExist:
        return JsonResponse({'error': 'One or more members not found'}, status=404)

def export_to_excel(request, record_id):
    from qc.views import format_date_indonesian, get_hari_indonesia
    import json
    
    try:
        record = BastRecordModel.objects.get(id=record_id)
    except BastRecordModel.DoesNotExist:
        return HttpResponse(status=404)

    file_path = os.path.join(os.path.dirname(__file__), 'static/bast/BAST.xlsx')
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    sheet.title = 'BAST'

    populate_bast_sheet(sheet, record)

    # Save the workbook to a BytesIO object
    def simplify_bast_id(bast_id):
        import re
        return re.sub(r'-(\d)([DPSM])$', r'-\2', bast_id)
    simple_bast_id = simplify_bast_id(record.bast_id)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={simple_bast_id}.xlsx'
    workbook.save(response)

    return response

def export_to_pdf(request, record_id):
    from qc.views import format_date_indonesian, get_hari_indonesia
    import json
    
    try:
        record = BastRecordModel.objects.get(id=record_id)
    except BastRecordModel.DoesNotExist:
        return HttpResponse(status=404)

    file_path = os.path.join(os.path.dirname(__file__), 'static/bast/BAST.xlsx')
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    sheet.title = 'BAST'

    populate_bast_sheet(sheet, record)

    # temporarily save the workbook to a file
    def simplify_bast_id(bast_id):
        import re
        return re.sub(r'-(\d)([DPSM])$', r'-\2', bast_id)
    simple_bast_id = simplify_bast_id(record.bast_id)
    temp_xlsx = os.path.join(os.path.dirname(__file__), f'static/bast/{simple_bast_id}.xlsx')
    workbook.save(temp_xlsx)
    temp_pdf_dir = os.path.join(os.path.dirname(__file__), 'static/bast')
    temp_pdf = os.path.join(temp_pdf_dir, f'{simple_bast_id}.pdf')

    import subprocess
    import sys
    
    # Try different possible LibreOffice executable names and paths
    libreoffice_paths = [
        'soffice',  # Linux/macOS
        'libreoffice',  # Some Linux distros
        'C:\\Program Files\\LibreOffice\\program\\soffice.exe',  # Windows default
        'C:\\Program Files (x86)\\LibreOffice\\program\\soffice.exe',  # 32-bit on 64-bit Windows
    ]
    
    success = False
    for soffice_cmd in libreoffice_paths:
        try:
            command = [soffice_cmd, '--headless', '--convert-to', 'pdf:calc_pdf_Export', 
                     temp_xlsx, '--outdir', temp_pdf_dir]
            result = subprocess.run(command, check=True, capture_output=True, text=True)
            if result.returncode == 0:
                success = True
                break
        except (subprocess.CalledProcessError, FileNotFoundError):
            continue
    
    if not success:
        # If we get here, none of the LibreOffice paths worked
        error_msg = (
            "PDF conversion failed: LibreOffice is not installed or not in PATH.\n"
            "Please install LibreOffice or add it to your system's PATH.\n"
            "You can download it from: https://www.libreoffice.org/\n"
            "After installation, try again."
        )
        return HttpResponse(error_msg, status=500, content_type='text/plain')
    
    # Clean up the temporary XLSX file
    if os.path.exists(temp_xlsx):
        try:
            os.remove(temp_xlsx)
        except Exception as e:
            print(f"Warning: Could not remove temporary file {temp_xlsx}: {e}")

    # Read the generated PDF file and return it in the response
    with open(temp_pdf, 'rb') as pdf_file:
        response = HttpResponse(pdf_file.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename={simple_bast_id}.pdf'

    if os.path.exists(temp_pdf):
        os.remove(temp_pdf)

    return response

def convert_to_roman(number):
    number = int(number)
    val = [
        1000, 900, 500, 400,
        100, 90, 50, 40,
        10, 9, 5, 4,
        1
        ]
    syb = [
        "M", "CM", "D", "CD",
        "C", "XC", "L", "XL",
        "X", "IX", "V", "IV",
        "I"
        ]
    roman_num = ''
    i = 0
    while  number > 0:
        for _ in range(number // val[i]):
            roman_num += syb[i]
            number -= val[i]
        i += 1
    return roman_num

def convert_to_indonesian(number):
    number = int(number)
    indonesian_numbers = ["Nol", "Satu", "Dua", "Tiga", "Empat", "Lima", "Enam", "Tujuh", "Delapan", "Sembilan", "Sepuluh"]
    if 0 <= number < len(indonesian_numbers):
        return indonesian_numbers[number]
    else:
        return str(number)

def get_cs_data(request, cs_id):
    try:
        cs_record = CsRecordModel.objects.get(cs_id=cs_id)
        return JsonResponse({
            'count_gaps': cs_record.count_gaps,
            'count_spikes': cs_record.count_spikes,
            'count_blanks': cs_record.count_blanks,
            'waktu_cs': cs_record.jam_pelaksanaan,
        })
    except CsRecordModel.DoesNotExist:
        return JsonResponse({'error': 'CS record not found'}, status=404)

def get_previous_poco_exp(request):
    """
    Returns the poco_exp value from the most recent BAST record.
    Returns None if no records exist.
    """
    try:
        latest_record = BastRecordModel.objects.latest('date')
        return JsonResponse({'poco_exp': latest_record.poco_exp})
    except BastRecordModel.DoesNotExist:
        return JsonResponse({'poco_exp': None}, status=404)


def get_previous_samsung_exp(request):
    """
    Returns the samsung_exp value from the most recent BAST record.
    Returns None if no records exist.
    """
    try:
        latest_record = BastRecordModel.objects.latest('date')
        return JsonResponse({'samsung_exp': latest_record.samsung_exp})
    except BastRecordModel.DoesNotExist:
        return JsonResponse({'samsung_exp': None}, status=404)


def get_previous_pulsa_poco(request):
    """
    Returns the pulsa_poco value from the most recent BAST record.
    Returns None if no records exist.
    """
    try:
        latest_record = BastRecordModel.objects.latest('date')
        return JsonResponse({'pulsa_poco': latest_record.pulsa_poco})
    except BastRecordModel.DoesNotExist:
        return JsonResponse({'pulsa_poco': None}, status=404)

def get_previous_members(request):
    """
    Returns the members data from the most recent BAST record.
    The response format is a list of dicts with 'nama' and 'keterangan' keys.
    """
    try:
        # Get the most recent BAST record
        latest_record = BastRecordModel.objects.latest('date', 'id')
        
        # Parse the member data (assuming it's stored as a JSON string in the member field)
        try:
            import json
            members = json.loads(latest_record.member) if latest_record.member else []
            return JsonResponse(members, safe=False)
        except json.JSONDecodeError:
            # If member data is not valid JSON, try to parse it as a simple string
            members = []
            if latest_record.member:
                # Assuming each line is a member name, and there's no keterangan
                for i, name in enumerate(latest_record.member.split('\n')):
                    if name.strip():
                        members.append({
                            'nama': name.strip(),
                            'keterangan': ''
                        })
            return JsonResponse(members, safe=False)
            
    except BastRecordModel.DoesNotExist:
        return JsonResponse([], safe=False)  # Return empty list if no records exist
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def populate_bast_sheet(sheet, record):
    from qc.views import format_date_indonesian, get_hari_indonesia
    import json
    import re

    member = json.loads(record.member)
    tanggal = format_date_indonesian(record.bast_id[5:-3])
    hari = get_hari_indonesia(record.bast_id[5:-3])

    sheet['J4'] = f'{convert_to_roman(record.kelompok)} ({convert_to_indonesian(record.kelompok)})'
    sheet['J6'] = f'{convert_to_roman(record.kel_berikut)} ({convert_to_indonesian(record.kel_berikut)})'
    sheet['N4'] = f': {tanggal}' 
    sheet['N5'] = f': {hari}'
    count_member = 0
    member_number = 1
    for idx, member_data in enumerate(member[:10]):  # Limit to 10 members to fit in the cells K9:K18 and L9:L18
        sheet[f'J{9 + idx}'] = member_number
        member_number += 1
        sheet[f'K{9 + idx}'] = member_data['nama']
        sheet[f'L{9 + idx}'] = member_data['keterangan']
        if re.match(r'^\s*hadir\s*$', member_data['keterangan'], re.IGNORECASE) or re.match(r'^\s*diganti\b.*$', member_data['keterangan'], re.IGNORECASE):
            count_member += 1

    sheet['L19'] = f'{count_member}'
    sheet['N6'] = f': {record.waktu_pelaksanaan}'
    sheet['G22'] = f'{record.event_indonesia}'
    sheet['G23'] = f'{record.event_luar}'
    sheet['G24'] = f'{record.event_indonesia + record.event_luar}'
    sheet['L22'] = f': {record.event_dirasakan} event'
    sheet['L23'] = f': {record.event_dikirim} event'
    sheet['E33'] = f'Pukul: {record.waktu_cs}'
    sheet['E34'] = f'IA (549) : Gaps = {record.count_gaps} ; Spike = {record.count_spikes} ; Blank = {record.count_blanks}'
    sheet['E38'] = f'Rp {record.pulsa_poco:,.0f}'.replace(',', '.')
    sheet['E40'] = f'{record.poco_exp.strftime("%d %b %Y")}'
    sheet['G40'] = f'{record.samsung_exp.strftime("%d %b %Y")}'
    sheet['C47'] = f'Jakarta, {tanggal}'
    sheet['C55'] = f'{record.spv}'
    sheet['C56'] = f'NIP. {record.NIP}'
    sheet['D44'] = f'{record.notes}'

        # import the events from the record using pandas
    events = pd.read_csv(StringIO(record.events))

    # add rows to the sheet
    rows_to_add = len(events)
    sheet.insert_rows(29, amount=rows_to_add)
    events = dataframe_to_rows(events, index=False, header=False)
    
    # insert the events to the sheet
    for r_idx, row in enumerate(events, 1):
        for c_idx, value in enumerate(row, 1):
            sheet.cell(row=r_idx+28, column=c_idx+2, value=value).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            # set the border of the first column to the left and the last column to the right, to thick
            sheet.cell(row=r_idx+28, column=2).border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='medium'))
            sheet.cell(row=r_idx+28, column=17).border = openpyxl.styles.Border(right=openpyxl.styles.Side(style='medium'))
            sheet.cell(row=r_idx+28, column=11).alignment = openpyxl.styles.Alignment(horizontal='left', vertical='center')
    
    # set the inserted cell border expanded to column 17 to thin
    default_row_height = 15.75
    for r_idx in range(rows_to_add):
        for c_idx in range(14):  # Iterate up to column 17 (index 14)
            cell = sheet.cell(row=r_idx + 29, column=c_idx + 3) # Get the cell object

            # Check the length of the value in index 12
            MMI_value = sheet.cell(row=r_idx + 29, column=12).value
            if pd.notna(MMI_value):
                if len(MMI_value) > 23:
                    # Calculate the new row height
                    new_height = default_row_height * ((len(MMI_value) // 23) + 1)
                    sheet.row_dimensions[r_idx + 29].height = new_height
                # Set the cell format to wrap text and center horizontally
                sheet.cell(row=r_idx + 29, column=12).alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='center')
            else:
                sheet.row_dimensions[r_idx + 29].height = default_row_height
                
            cell.border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style='thin'),
                right=openpyxl.styles.Side(style='thin'),
                top=openpyxl.styles.Side(style='thin'),
                bottom=openpyxl.styles.Side(style='thin')
            )


def export_bast_to_csv(request):
    """
    Export all BAST records to a CSV file.
    """
    try:
        # Log the start of the export
        print("Starting BAST records export...")
        
        # Create the HttpResponse object with the appropriate CSV header
        response = HttpResponse(
            content_type='text/csv',
            headers={
                'Content-Disposition': 'attachment; filename="bast_records_export.csv"',
                'Cache-Control': 'no-cache, no-store, must-revalidate',
                'Pragma': 'no-cache',
                'Expires': '0'
            },
        )
        
        # Force the response to be treated as a download
        response['Content-Encoding'] = 'UTF-8'
        response['Content-Type'] = 'text/csv; charset=utf-8-sig'  # Add BOM for Excel
        
        # Create a CSV writer with proper quoting
        writer = csv.writer(response, quoting=csv.QUOTE_ALL, delimiter=',')
        
        # Write UTF-8 BOM for Excel compatibility
        response.write('\ufeff')
        
        # Write headers
        writer.writerow([
            'BAST ID', 'Date', 'Waktu Pelaksanaan', 'Shift', 'Kelompok',
            'Kelompok Berikut', 'Events', 'Supervisor', 'NIP', 'Event Indonesia',
            'Event Luar', 'Event Dirasakan', 'Event Dikirim', 'Members',
            'Count Gaps', 'Count Spikes', 'Count Blanks', 'Waktu CS',
            'Pulsa Poco', 'POCO Expiry', 'Samsung Expiry', 'Notes'
        ])
        
        # Get all records ordered by bast_id
        records = BastRecordModel.objects.all().order_by('bast_id')
        print(f"Found {records.count()} records to export")
        
        # Write data rows
        for record in records:
            try:
                writer.writerow([
                    record.bast_id or '',
                    record.date.strftime('%Y-%m-%d') if record.date else '',
                    str(record.waktu_pelaksanaan) if record.waktu_pelaksanaan else '',
                    str(record.shift) if record.shift else '',
                    str(record.kelompok) if record.kelompok else '',
                    str(record.kel_berikut) if record.kel_berikut else '',
                    record.events or '',
                    str(record.spv) if record.spv else '',
                    record.NIP or '',
                    record.event_indonesia or 0,
                    record.event_luar or 0,
                    record.event_dirasakan or 0,
                    record.event_dikirim or 0,
                    record.member or '',
                    record.count_gaps or 0,
                    record.count_spikes or 0,
                    record.count_blanks or 0,
                    record.waktu_cs or '',
                    record.pulsa_poco or 0,
                    record.poco_exp.strftime('%Y-%m-%d') if record.poco_exp else '',
                    record.samsung_exp.strftime('%Y-%m-%d') if record.samsung_exp else '',
                    record.notes or ''
                ])
            except Exception as e:
                print(f"Error writing record {getattr(record, 'bast_id', 'unknown')}: {str(e)}")
                # Skip the problematic record and continue with the next one
                continue
        
        print("Export completed successfully")
        return response
        
    except Exception as e:
        print(f"Error in export_bast_to_csv: {str(e)}")
        # Return an error response
        return HttpResponse(
            f"Error generating CSV: {str(e)}",
            status=500,
            content_type='text/plain'
        )